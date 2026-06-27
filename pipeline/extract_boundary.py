"""
Extract RESEX boundary from shapefile to GeoJSON + generate meta.json with stats.
"""

import json
import geopandas as gpd
import openpyxl
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent
WEB_ASSETS = DATA_DIR / "web_assets"


def extract_geojson():
    """Convert shapefile to simplified GeoJSON."""
    gdf = gpd.read_file(DATA_DIR / "RESEX_Jaci_parana" / "ucs.shp")
    gdf = gdf.to_crs(epsg=4326)
    # Simplify geometry (~100m tolerance) for web
    gdf_simple = gdf.copy()
    gdf_simple.geometry = gdf.geometry.simplify(0.001)
    gdf_simple.to_file(WEB_ASSETS / "resex_boundary.geojson", driver="GeoJSON")

    bounds = gdf.total_bounds  # [minx, miny, maxx, maxy]
    centroid = gdf.geometry.centroid.iloc[0]
    print(f"Boundary: {bounds}")
    print(f"Centroid: ({centroid.x:.4f}, {centroid.y:.4f})")
    return bounds, centroid


def extract_stats():
    """Extract yearly hectare stats from Excel."""
    wb = openpyxl.load_workbook(DATA_DIR / "evolucao_cobertura.xlsx")
    ws = wb.active

    # Read header row to find year columns
    headers = [cell.value for cell in ws[1]]
    year_start_col = None
    for i, h in enumerate(headers):
        if isinstance(h, int) and h >= 1985:
            year_start_col = i
            break

    years_available = [h for h in headers[year_start_col:] if isinstance(h, int)]

    # Read class data
    stats = {}
    class_names = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        class_code = row[1]
        class_name = row[3]  # class_level_1
        class_names[class_code] = class_name
        for i, year in enumerate(years_available):
            col_idx = year_start_col + i
            val = row[col_idx]
            if val is not None:
                if year not in stats:
                    stats[year] = {}
                stats[year][str(class_code)] = round(float(val), 1)

    return stats, class_names, years_available


def build_meta(bounds, centroid, stats, class_names, years_available):
    """Build meta.json with all context data."""
    import glob

    # Get TIFF years
    tiff_files = sorted(Path(DATA_DIR / "cobertura_tiff" / "tiff").glob("cobertua_*.tif"))
    tiff_years = [int(f.stem.split("_")[-1]) for f in tiff_files]

    # Build yearly summary (forest + pasture hectares)
    yearly_summary = {}
    for year in years_available:
        if year in stats:
            s = stats[year]
            forest_ha = s.get("3", 0) + s.get("6", 0)  # forest + floodable
            pasture_ha = s.get("15", 0)
            yearly_summary[str(year)] = {
                "forest_ha": round(forest_ha, 1),
                "pasture_ha": round(pasture_ha, 1),
                "total_ha": round(sum(s.values()), 1),
            }

    meta = {
        "bbox": [round(bounds[0], 5), round(bounds[1], 5),
                 round(bounds[2], 5), round(bounds[3], 5)],
        "center": [round(centroid.x, 5), round(centroid.y, 5)],
        "tiff_years": tiff_years,
        "all_years": years_available,
        "raster_size": [1767, 2923],
        "class_names": {str(k): v for k, v in class_names.items()},
        "stats": yearly_summary,
    }

    with open(WEB_ASSETS / "meta.json", "w") as f:
        json.dump(meta, f, indent=2)
    print(f"Saved meta.json ({len(yearly_summary)} yearly entries)")
    return meta


def run():
    WEB_ASSETS.mkdir(parents=True, exist_ok=True)
    bounds, centroid = extract_geojson()
    stats, class_names, years = extract_stats()
    meta = build_meta(bounds, centroid, stats, class_names, years)
    return meta


if __name__ == "__main__":
    run()
